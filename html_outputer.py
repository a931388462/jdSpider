import time
from openpyxl import Workbook, load_workbook


class HtmlOutputer(object):
    def __init__(self):
        self.pages_commodity = []

    def collect_data(self, commoditys):
        if commoditys is None:
            return
        self.pages_commodity.append(commoditys)

    def output_html(self,craw_comm,crawled_list):
        dateStr = time.strftime("%Y%m%d", time.localtime())
        fileName = craw_comm + "的爬取结果_" + dateStr + ".xlsx"
        # 文件不存在的情况，创建文件
        try:
            wb = load_workbook(fileName)
        except:
            # 创建文件对象
            wb = Workbook()
            # 获取第一个sheet
            ws = wb.active
            # 表头
            ws['A1'] = "商品ID"
            ws['B1'] = "商品名"
            ws['C1'] = "价格"
            ws['D1'] = "选购指数"
            ws['E1'] = "店铺名"
            # 调整列宽
            ws.column_dimensions['A'].width = 15.0
            # 调整列宽
            ws.column_dimensions['B'].width = 100.0
            # 调整列宽
            ws.column_dimensions['E'].width = 50.0
        # 获取第一个sheet
        ws = wb.active
        #写入时间
        ws.append(["--------------------------------------------------------------------------------------"])
        for commoditys in self.pages_commodity:
                for commodity in commoditys:
                    ss = crawled_list[craw_comm]
                    if commodity.commID not in ss:
                        # 添加到已爬取列表中
                        ss.add(commodity.commID)
                        # 写入多个单元格
                        ws.append([commodity.commID, commodity.commName, commodity.price, commodity.purchasingIndex
                                      , commodity.shopName])

        # 保存为爬取结果
        wb.save(fileName)